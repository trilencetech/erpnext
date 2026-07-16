# Copyright (c) 2026, TriLence Tech and contributors
# Purchase Receipt Register — Script Report

from collections import OrderedDict
from datetime import datetime
import json

import frappe
from frappe import _

from erpnext.utilities.report_style import (
    COLORS,
    KIND_COLORS,
    badge_html,
    cell_fill,
    date_range_label,
    fmt_date,
    merge_and_style,
    render_print_shell,
    render_stats_bar,
    style_cell,
    write_row,
)

# Purchase Receipt `status` -> stats-bar / badge "kind"
STATUS_KIND = {
    "Completed": "ok",
    "To Bill": "warn",
    "Partly Billed": "warn",
    "Return Issued": "error",
    "Closed": "info",
}


def execute(filters=None):
    columns = get_columns()
    rows, _stats = _get_report_data(filters or {})
    return columns, rows


# ── COLUMNS ───────────────────────────────────────────────────────────────────

def get_columns():
    return [
        {
            "label": _("Item Code"), "fieldname": "item_code",
            "fieldtype": "Link", "options": "Item", "width": 150,
        },
        {
            "label": _("Item Group / Supplier"), "fieldname": "item_group",
            "fieldtype": "Data", "width": 160,
        },
        {
            "label": _("Purchase Receipt"), "fieldname": "purchase_receipt",
            "fieldtype": "Link", "options": "Purchase Receipt", "width": 160,
        },
        {
            "label": _("Date"), "fieldname": "posting_date",
            "fieldtype": "Date", "width": 100,
        },
        {
            "label": _("Qty"), "fieldname": "qty",
            "fieldtype": "Float", "width": 90, "precision": 3,
        },
        {
            "label": _("Rate"), "fieldname": "rate",
            "fieldtype": "Currency", "width": 100,
        },
        {
            "label": _("Amount"), "fieldname": "amount",
            "fieldtype": "Currency", "width": 120,
        },
        {
            "label": _("Warehouse"), "fieldname": "warehouse",
            "fieldtype": "Link", "options": "Warehouse", "width": 160,
        },
        {
            "label": _("Status"), "fieldname": "status",
            "fieldtype": "Data", "width": 160,
        },
    ]


# ── DATA ──────────────────────────────────────────────────────────────────────

def _get_report_data(filters):
    """Return (rows, stats) for the given filters — shared by the grid, print and Excel views."""
    conditions = ["pr.docstatus = 1"]
    values = {}

    if filters.get("company"):
        conditions.append("pr.company = %(company)s")
        values["company"] = filters["company"]

    if filters.get("from_date"):
        conditions.append("pr.posting_date >= %(from_date)s")
        values["from_date"] = filters["from_date"]

    if filters.get("to_date"):
        conditions.append("pr.posting_date <= %(to_date)s")
        values["to_date"] = filters["to_date"]

    if filters.get("supplier"):
        conditions.append("pr.supplier = %(supplier)s")
        values["supplier"] = filters["supplier"]

    where = "WHERE " + " AND ".join(conditions)

    rows = frappe.db.sql(
        f"""
        SELECT
            pr.name          AS purchase_receipt,
            pr.supplier       AS supplier,
            pr.posting_date   AS posting_date,
            pr.status         AS status,
            pri.item_code     AS item_code,
            pri.item_group    AS item_group,
            pri.qty           AS qty,
            pri.rate          AS rate,
            pri.amount        AS amount,
            pri.warehouse     AS warehouse,
            pri.rejected_qty  AS rejected_qty
        FROM `tabPurchase Receipt Item` pri
        JOIN `tabPurchase Receipt` pr ON pr.name = pri.parent
        {where}
        ORDER BY pr.posting_date, pr.name, pri.idx
        """,
        values,
        as_dict=True,
    )

    return _build_grouped_rows(rows)


def _build_grouped_rows(raw_rows):
    docs = OrderedDict()
    for r in raw_rows:
        pr = r.purchase_receipt
        if pr not in docs:
            docs[pr] = {
                "supplier":     r.supplier or "",
                "posting_date": r.posting_date,
                "status":       r.status or "",
                "items":        [],
            }
        docs[pr]["items"].append(r)

    result      = []
    stat_counts = {"ok": 0, "warn": 0, "error": 0, "info": 0, "neutral": 0}
    grand_qty   = 0.0
    grand_amount = 0.0

    for pr, info in docs.items():
        kind = STATUS_KIND.get(info["status"], "neutral")
        stat_counts[kind] += 1

        # Group header — one Purchase Receipt voucher
        result.append({
            "item_code":        "",
            "item_group":       info["supplier"],
            "purchase_receipt": pr,
            "posting_date":     info["posting_date"],
            "qty":              None,
            "rate":             None,
            "amount":           None,
            "warehouse":        "",
            "status":           info["status"],
            "status_kind":      kind,
            "rejected_qty":     0,
            "row_type":         "group_header",
            "bold":             1,
        })

        sub_qty    = 0.0
        sub_amount = 0.0

        for it in info["items"]:
            qty    = round(float(it.qty or 0), 3)
            amount = round(float(it.amount or 0), 2)
            sub_qty    += qty
            sub_amount += amount

            result.append({
                "item_code":        it.item_code or "",
                "item_group":       it.item_group or "",
                "purchase_receipt": "",
                "posting_date":     None,
                "qty":              qty,
                "rate":             round(float(it.rate or 0), 2),
                "amount":           amount,
                "warehouse":        it.warehouse or "",
                "status":           "",
                "status_kind":      "",
                "rejected_qty":     round(float(it.rejected_qty or 0), 3),
                "row_type":         "data",
            })

        # Subtotal per Purchase Receipt
        result.append({
            "item_code":        "",
            "item_group":       "",
            "purchase_receipt": f"{len(info['items'])} item(s)",
            "posting_date":     None,
            "qty":              round(sub_qty, 3),
            "rate":             None,
            "amount":           round(sub_amount, 2),
            "warehouse":        "",
            "status":           "",
            "status_kind":      kind,
            "rejected_qty":     0,
            "row_type":         "subtotal",
            "bold":             1,
        })

        grand_qty    += sub_qty
        grand_amount += sub_amount

    # Grand total
    result.append({
        "item_code":        "Grand Total",
        "item_group":       f"{len(docs)} receipt(s)",
        "purchase_receipt": "",
        "posting_date":     None,
        "qty":              round(grand_qty, 3),
        "rate":             None,
        "amount":           round(grand_amount, 2),
        "warehouse":        "",
        "status":           "",
        "status_kind":      "",
        "rejected_qty":     0,
        "row_type":         "grand_total",
        "bold":             1,
    })

    stats = [
        {"label": _("Total Receipts"),  "value": len(docs),            "kind": "neutral"},
        {"label": _("Completed"),       "value": stat_counts["ok"],    "kind": "ok"},
        {"label": _("Pending Billing"), "value": stat_counts["warn"],  "kind": "warn"},
        {"label": _("Returned"),        "value": stat_counts["error"], "kind": "error"},
        {"label": _("Other"),           "value": stat_counts["info"],  "kind": "info"},
    ]

    return result, stats


# ── PRINT ─────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_print_html(filters):
    if isinstance(filters, str):
        filters = json.loads(filters)

    rows, stats = _get_report_data(filters)
    table_html  = _build_table_html(rows)
    stats_html  = render_stats_bar(stats)
    subtitle    = date_range_label(filters.get("from_date"), filters.get("to_date"))

    return render_print_shell(
        "Purchase Receipt Register", filters.get("company", ""), subtitle, stats_html, table_html
    )


def _build_table_html(rows):
    html = """
<table>
  <thead>
    <tr>
      <th>ITEM CODE</th>
      <th>ITEM GROUP / SUPPLIER</th>
      <th>PURCHASE RECEIPT</th>
      <th>DATE</th>
      <th class="r">QTY</th>
      <th class="r">RATE</th>
      <th class="r">AMOUNT</th>
      <th>WAREHOUSE</th>
      <th>STATUS</th>
    </tr>
  </thead>
  <tbody>
"""

    for row in rows:
        rtype = row["row_type"]

        if rtype == "group_header":
            kind = row["status_kind"]
            html += (
                f'<tr class="grp-{kind}">'
                f'<td></td>'
                f'<td><strong>{row["item_group"]}</strong></td>'
                f'<td><strong>{row["purchase_receipt"]}</strong></td>'
                f'<td>{fmt_date(row["posting_date"])}</td>'
                f'<td></td><td></td><td></td><td></td>'
                f'<td>{badge_html(row["status"], kind)}</td>'
                f'</tr>\n'
            )

        elif rtype == "data":
            status_cell = ""
            if row["rejected_qty"]:
                status_cell = badge_html(f'Rejected {row["rejected_qty"]:.3f}', "error")
            html += (
                f'<tr>'
                f'<td>{row["item_code"]}</td>'
                f'<td>{row["item_group"]}</td>'
                f'<td></td><td></td>'
                f'<td class="r">{row["qty"]:.3f}</td>'
                f'<td class="r">{row["rate"]:.2f}</td>'
                f'<td class="r">{row["amount"]:.2f}</td>'
                f'<td>{row["warehouse"]}</td>'
                f'<td>{status_cell}</td>'
                f'</tr>\n'
            )

        elif rtype == "subtotal":
            kind = row["status_kind"]
            html += (
                f'<tr class="sub-{kind}">'
                f'<td colspan="3" style="text-align:right;padding-right:8px">{row["purchase_receipt"]}</td>'
                f'<td></td>'
                f'<td class="r">{row["qty"]:.3f}</td>'
                f'<td></td>'
                f'<td class="r">{row["amount"]:.2f}</td>'
                f'<td></td><td></td>'
                f'</tr>\n'
            )

        elif rtype == "grand_total":
            html += (
                f'<tr class="grand-total">'
                f'<td><strong>{row["item_code"]}</strong></td>'
                f'<td colspan="2">{row["item_group"]}</td>'
                f'<td></td>'
                f'<td class="r"><strong>{row["qty"]:.3f}</strong></td>'
                f'<td></td>'
                f'<td class="r"><strong>{row["amount"]:.2f}</strong></td>'
                f'<td></td><td></td>'
                f'</tr>\n'
            )

    return html + "</tbody></table>"


# ── EXCEL ─────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_excel(filters):
    import base64
    import io

    try:
        import openpyxl
    except ImportError:
        frappe.throw(_("openpyxl not installed. Run: <code>bench pip install openpyxl</code>"))

    from openpyxl.utils import get_column_letter

    if isinstance(filters, str):
        filters = json.loads(filters)

    rows, stats = _get_report_data(filters)
    company  = filters.get("company", "")
    subtitle = date_range_label(filters.get("from_date"), filters.get("to_date"))
    gen_dt   = datetime.now().strftime("%d-%m-%Y %H:%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Receipt Register"

    COLS = 9
    for i, w in enumerate([16, 22, 18, 12, 10, 12, 14, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    merge_and_style(
        ws, r, COLS, f"{company}  —  Purchase Receipt Register  —  {subtitle}",
        COLORS["header_bg"], COLORS["header_fg"], bold=True, size=13,
    )
    ws.row_dimensions[r].height = 22
    r += 1

    merge_and_style(ws, r, COLS, f"Generated: {gen_dt}", COLORS["neutral_bg"], "555555", size=9, h="left")
    ws.row_dimensions[r].height = 14
    r += 1

    # Stats row
    for col, s in enumerate(stats, 1):
        bg, fg, _b = KIND_COLORS.get(s["kind"], KIND_COLORS["neutral"])
        style_cell(ws.cell(row=r, column=col, value=f"{s['label']}: {s['value']}"), bg, fg=fg, bold=True, size=10, h="center")
    for col in range(len(stats) + 1, COLS + 1):
        ws.cell(row=r, column=col).fill = cell_fill("FFFFFF")
    ws.row_dimensions[r].height = 18
    r += 1

    # Spacer
    for col in range(1, COLS + 1):
        ws.cell(row=r, column=col).fill = cell_fill("FFFFFF")
    ws.row_dimensions[r].height = 6
    r += 1

    # Column headers
    hdrs = ["ITEM CODE", "ITEM GROUP / SUPPLIER", "PURCHASE RECEIPT", "DATE",
            "QTY", "RATE", "AMOUNT", "WAREHOUSE", "STATUS"]
    for col, h in enumerate(hdrs, 1):
        style_cell(
            ws.cell(row=r, column=col, value=h),
            COLORS["header_bg"], fg=COLORS["header_fg"], bold=True, size=10,
            h="right" if col in (5, 6, 7) else "left",
        )
    ws.row_dimensions[r].height = 18
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    r += 1

    haligns = ["left", "left", "left", "left", "right", "right", "right", "left", "left"]

    for row in rows:
        rtype = row["row_type"]

        if rtype == "group_header":
            bg, fg, _b = KIND_COLORS.get(row["status_kind"], KIND_COLORS["neutral"])
            write_row(
                ws, r,
                ["", row["item_group"], row["purchase_receipt"], fmt_date(row["posting_date"]), "", "", "", "", row["status"]],
                bg, fgs=[fg] * 9,
                bolds=[False, True, True, False, False, False, False, False, True],
                haligns=haligns, row_h=18, top_border=True,
            )

        elif rtype == "data":
            status_txt = f'Rejected {row["rejected_qty"]:.3f}' if row["rejected_qty"] else ""
            write_row(
                ws, r,
                [row["item_code"], row["item_group"], "", "", f'{row["qty"]:.3f}',
                 f'{row["rate"]:.2f}', f'{row["amount"]:.2f}', row["warehouse"], status_txt],
                "FFFFFF",
                fgs=["111111"] * 8 + ["991B1B" if status_txt else "111111"],
                bolds=[False] * 8 + [bool(status_txt)],
                haligns=haligns,
            )

        elif rtype == "subtotal":
            bg, fg, _b = KIND_COLORS.get(row["status_kind"], KIND_COLORS["neutral"])
            write_row(
                ws, r,
                ["", "", row["purchase_receipt"], "", f'{row["qty"]:.3f}', "", f'{row["amount"]:.2f}', "", ""],
                bg, fgs=[fg] * 9, bolds=[True] * 9, haligns=haligns,
            )

        elif rtype == "grand_total":
            write_row(
                ws, r,
                [row["item_code"], row["item_group"], "", "", f'{row["qty"]:.3f}', "", f'{row["amount"]:.2f}', "", ""],
                COLORS["header_bg"], fgs=[COLORS["header_fg"]] * 9, bolds=[True] * 9,
                haligns=haligns, row_h=18, top_border=True,
            )

        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return {
        "data":     base64.b64encode(buf.getvalue()).decode(),
        "filename": "purchase_receipt_register.xlsx",
    }

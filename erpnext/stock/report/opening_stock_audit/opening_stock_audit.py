from collections import OrderedDict
from frappe import _
import frappe
import json


def execute(filters=None):
    columns = get_columns()
    data    = get_data(filters)
    return columns, data


# ── COLUMNS ───────────────────────────────────────────────────────────────────

def get_columns():
    return [
        {
            "label": _("Item Code"), "fieldname": "item_code",
            "fieldtype": "Link", "options": "Item", "width": 180,
        },
        {
            "label": _("Item Group"), "fieldname": "item_group",
            "fieldtype": "Data", "width": 150,
        },
        {
            "label": _("Stock Entry"), "fieldname": "stock_entry",
            "fieldtype": "Link", "options": "Stock Entry", "width": 180,
        },
        {
            "label": _("Date"), "fieldname": "posting_date",
            "fieldtype": "Date", "width": 100,
        },
        {
            "label": _("Qty (Kgs)"), "fieldname": "qty",
            "fieldtype": "Float", "width": 120, "precision": 3,
        },
        {
            "label": _("Warehouse"), "fieldname": "warehouse",
            "fieldtype": "Link", "options": "Warehouse", "width": 200,
        },
        {
            "label": _("Sold"), "fieldname": "sold",
            "fieldtype": "Data", "width": 80,
        },
    ]


# ── DATA ──────────────────────────────────────────────────────────────────────

def get_data(filters):
    conditions = ["se.docstatus = 1"]
    values     = {}

    if filters.get("company"):
        conditions.append("se.company = %(company)s")
        values["company"] = filters["company"]

    if filters.get("from_date"):
        conditions.append("se.posting_date >= %(from_date)s")
        values["from_date"] = filters["from_date"]

    if filters.get("to_date"):
        conditions.append("se.posting_date <= %(to_date)s")
        values["to_date"] = filters["to_date"]

    if filters.get("stock_entry_type"):
        conditions.append("se.stock_entry_type = %(stock_entry_type)s")
        values["stock_entry_type"] = filters["stock_entry_type"]

    where = "WHERE " + " AND ".join(conditions)

    rows = frappe.db.sql(
        f"""
        SELECT
            sed.item_code                           AS item_code,
            sed.item_id                             AS item_id,
            item.item_group                         AS item_group,
            sed.parent                              AS stock_entry,
            se.posting_date                         AS posting_date,
            sed.qty                                 AS qty,
            COALESCE(sed.t_warehouse, sed.s_warehouse) AS warehouse
        FROM `tabStock Entry Detail` sed
        JOIN `tabStock Entry`  se   ON se.name   = sed.parent
        JOIN `tabItem`         item ON item.name  = sed.item_code
        {where}
        ORDER BY item.item_group, sed.item_code, se.posting_date, se.name
        """,
        values,
        as_dict=True,
    )

    item_ids  = list({r.item_id for r in rows if r.item_id})
    sold_ids  = _get_sold_item_ids(item_ids) if item_ids else set()

    return _build_grouped_rows(rows, sold_ids)


def _get_sold_item_ids(item_ids):
    """Return the set of item_ids that appear in any submitted Delivery Note."""
    rows = frappe.db.sql(
        """
        SELECT DISTINCT dni.item_id
        FROM `tabDelivery Note Item` dni
        JOIN `tabDelivery Note` dn ON dn.name = dni.parent
        WHERE dn.docstatus = 1
          AND dni.item_id IN %s
        """,
        (item_ids,),
        as_dict=True,
    )
    return {r.item_id for r in rows}


def _build_grouped_rows(raw_data, sold_ids=None):
    sold_ids = sold_ids or set()

    # ── Pass 1: group entries by item_code ────────────────────────────────────
    items = OrderedDict()   # {item_code: {"item_group": str, "entries": [...]}}

    for row in raw_data:
        code = row.item_code or ""
        if code not in items:
            items[code] = {"item_group": row.item_group or "", "entries": []}
        items[code]["entries"].append({
            "stock_entry":  row.stock_entry,
            "posting_date": row.posting_date,
            "qty":          round(float(row.qty or 0), 3),
            "warehouse":    row.warehouse or "",
            "item_id":      row.item_id or "",
            "is_sold":      1 if (row.item_id and row.item_id in sold_ids) else 0,
        })

    # ── Pass 2: build display rows ────────────────────────────────────────────
    result      = []
    grand_qty   = 0.0
    grand_count = 0

    for code, info in items.items():
        entries     = info["entries"]
        entry_count = len(entries)
        total_qty   = round(sum(e["qty"] for e in entries), 3)
        # Group is flagged sold if ANY individual entry's item_id was found in DN
        group_sold  = 1 if any(e["is_sold"] for e in entries) else 0

        # Group header row — item_code acts as the clickable link
        result.append({
            "item_code":    code,
            "item_group":   info["item_group"],
            "stock_entry":  "",
            "posting_date": None,
            "qty":          None,
            "warehouse":    "",
            "sold":         "",
            "bold":         1,
            "row_type":     "group_header",
            "entry_count":  entry_count,
            "is_sold":      0,
        })

        # Individual stock entry lines — each carries its own item_id sold flag
        for entry in entries:
            result.append({
                "item_code":    "",
                "item_group":   "",
                "stock_entry":  entry["stock_entry"],
                "posting_date": entry["posting_date"],
                "qty":          entry["qty"],
                "warehouse":    entry["warehouse"],
                "sold":         "Yes" if entry["is_sold"] else "",
                "row_type":     "data",
                "entry_count":  entry_count,
                "is_sold":      entry["is_sold"],
            })

        # Subtotal per item_code
        result.append({
            "item_code":    "",
            "item_group":   "",
            "stock_entry":  f"{'⚠ Multiple entries — ' if entry_count > 1 else ''}{entry_count} {'entry' if entry_count == 1 else 'entries'}",
            "posting_date": None,
            "qty":          total_qty,
            "warehouse":    "",
            "sold":         "",
            "bold":         1,
            "row_type":     "subtotal",
            "entry_count":  entry_count,
            "is_sold":      0,
        })

        grand_qty   += total_qty
        grand_count += entry_count

    # Grand Total
    result.append({
        "item_code":    "Grand Total",
        "item_group":   f"{len(items)} item(s)  |  {grand_count} entries",
        "stock_entry":  "",
        "posting_date": None,
        "qty":          round(grand_qty, 3),
        "warehouse":    "",
        "sold":         "",
        "bold":         1,
        "row_type":     "grand_total",
    })

    return result


# ── PDF / Print ────────────────────────────────────────────────────────────────

@frappe.whitelist()
def pdf_to_excel(file_data, file_name=None):
    """Accept a base64-encoded PDF, extract its tables, and return a styled Excel file as base64."""
    import base64
    import io
    import re

    for lib, pkg in [("pdfplumber", "pdfplumber"), ("openpyxl", "openpyxl")]:
        try:
            __import__(lib)
        except ImportError:
            frappe.throw(
                _(f"{lib} is not installed. Run: <code>bench pip install {pkg}</code>")
            )

    import pdfplumber
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # ── Extract rows from all PDF tables ─────────────────────────────────────
    pdf_bytes = base64.b64decode(file_data)
    rows = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for tbl in (page.extract_tables() or []):
                for row in tbl:
                    cleaned = [str(c).strip() if c else "" for c in row]
                    if any(cleaned):
                        rows.append(cleaned)

    if not rows:
        frappe.throw(_("No table data found in this PDF."))

    # ── Style helpers ─────────────────────────────────────────────────────────
    _thin = Side(style="thin", color="CCCCCC")
    _bord = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="111111", size=10):
        return Font(bold=bold, color=color, name="Arial", size=size)

    def _style(ws_row, bg, fg="111111", bold=False, size=10, right_cols=()):
        for idx, cell in enumerate(ws_row, 1):
            cell.fill      = _fill(bg)
            cell.font      = _font(bold, fg, size)
            cell.border    = _bord
            cell.alignment = Alignment(
                horizontal="right" if idx in right_cols else "left",
                vertical="center",
            )

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opening Stock Audit"

    current_multi = False   # tracks whether the current item-group has multiple entries

    for row_idx, raw in enumerate(rows, 1):
        while len(raw) < 6:
            raw.append("")
        raw = raw[:6]

        ws.append(raw)
        ws_row = ws[row_idx]

        c0 = raw[0].upper().strip()
        c1 = raw[1]
        c2 = raw[2]

        # Column header row
        if c0 in ("ITEM CODE", "ITEM\nCODE"):
            _style(ws_row, "1A3A6A", fg="FFFFFF", bold=True, size=10, right_cols=(5,))
            ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)

        # Grand total
        elif c0 == "GRAND TOTAL":
            _style(ws_row, "1A3A6A", fg="FFFFFF", bold=True, size=11, right_cols=(5,))

        # Group header: col1 has text, cols 3–6 are blank
        elif raw[0] and not raw[2] and not raw[3] and not raw[4]:
            current_multi = "⚠" in c1 or (
                "entries" in c1.lower() and not c1.lower().startswith("✓")
            )
            if current_multi:
                _style(ws_row, "FEF3C7", fg="92400E", bold=True)
            else:
                _style(ws_row, "D1FAE5", fg="064E3B", bold=True)

        # Subtotal: cols 1–2 blank, col3 mentions "entr"
        elif not raw[0] and not raw[1] and "entr" in c2.lower():
            warn = "⚠" in c2 or current_multi
            _style(
                ws_row,
                "FDE68A" if warn else "DCFCE7",
                fg="92400E" if warn else "111111",
                bold=True, right_cols=(5,),
            )
            current_multi = False

        # Data row
        else:
            _style(ws_row, "FFFBEB" if current_multi else "F0FDF4", right_cols=(5,))

    # Column widths and row heights
    for i, w in enumerate([20, 18, 22, 13, 14, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    out_name = re.sub(r"\.pdf$", "", file_name or "opening_stock_audit", flags=re.I) + ".xlsx"
    return {"data": base64.b64encode(buf.getvalue()).decode(), "filename": out_name}


@frappe.whitelist()
def get_diff_report_html(file_data, filters):
    """Compare uploaded external Excel with system Opening Stock entries — HTML report."""
    import json
    if isinstance(filters, str):
        filters = json.loads(filters)
    try:
        __import__("openpyxl")
    except ImportError:
        frappe.throw(_("openpyxl not installed. Run: <code>bench pip install openpyxl</code>"))
    items_data, stats = _build_diff_data(file_data, filters)
    return _render_diff_html(filters, items_data, stats)


@frappe.whitelist()
def get_diff_excel(file_data, filters):
    """Compare uploaded external Excel with system Opening Stock entries — styled Excel download."""
    import base64, io, json
    from datetime import datetime

    if isinstance(filters, str):
        filters = json.loads(filters)
    try:
        import openpyxl as _xl
    except ImportError:
        frappe.throw(_("openpyxl not installed. Run: <code>bench pip install openpyxl</code>"))

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    items_data, stats = _build_diff_data(file_data, filters)

    company   = filters.get("company", "")
    from_date = str(filters.get("from_date", ""))
    to_date   = str(filters.get("to_date", ""))

    def fmt(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").strftime("%d-%m-%Y")
        except Exception:
            return d

    date_range = f"{fmt(from_date)} to {fmt(to_date)}" if from_date != to_date else fmt(from_date)
    gen_dt     = datetime.now().strftime("%d-%m-%Y %H:%M")

    # ── Style helpers ─────────────────────────────────────────────────────────
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="111111", size=10, italic=False):
        return Font(bold=bold, color=color, name="Arial", size=size, italic=italic)

    _thin     = Side(style="thin",   color="D0D0D0")
    _med      = Side(style="medium", color="888888")
    _bord     = Border(left=_thin, right=_thin, top=_thin,  bottom=_thin)
    _bord_top = Border(left=_thin, right=_thin, top=_med,   bottom=_thin)

    def _align(h="left", v="center"):
        return Alignment(horizontal=h, vertical=v, wrap_text=False)

    def _style(cell, bg, fg="111111", bold=False, size=10,
               h="left", italic=False, border=None):
        cell.fill      = _fill(bg)
        cell.font      = _font(bold=bold, color=fg, size=size, italic=italic)
        cell.alignment = _align(h)
        cell.border    = border or _bord

    # ── Workbook setup ────────────────────────────────────────────────────────
    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "Stock Diff Report"

    COLS = 7
    for i, w in enumerate([18, 13, 26, 13, 12, 15, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def _merge(row_num, value, bg, fg, bold=False, size=10, h="center"):
        ws.merge_cells(f"A{row_num}:{get_column_letter(COLS)}{row_num}")
        c = ws.cell(row=row_num, column=1, value=value)
        _style(c, bg, fg=fg, bold=bold, size=size, h=h)
        for col in range(2, COLS + 1):
            ws.cell(row=row_num, column=col).fill = _fill(bg)

    def _write(row_num, values, bg, fgs=None, bolds=None, haligns=None,
                sizes=None, italics=None, row_h=17, top_border=False):
        brd = _bord_top if top_border else _bord
        for col, val in enumerate(values, 1):
            c      = ws.cell(row=row_num, column=col, value=val)
            _style(c, bg,
                   fg      = fgs[col-1]    if fgs     else "111111",
                   bold    = bolds[col-1]  if bolds   else False,
                   h       = haligns[col-1] if haligns else "left",
                   size    = sizes[col-1]  if sizes   else 10,
                   italic  = italics[col-1] if italics else False,
                   border  = brd)
        ws.row_dimensions[row_num].height = row_h

    r = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    _merge(r, f"{company}  —  Stock Difference Report  —  {date_range}",
           "1A3A6A", "FFFFFF", bold=True, size=13)
    ws.row_dimensions[r].height = 22
    r += 1

    _merge(r, f"Generated: {gen_dt}   |   External Excel vs Opening Stock Audit",
           "F1F5F9", "555555", size=9)
    ws.row_dimensions[r].height = 14
    r += 1

    # ── Stats row ─────────────────────────────────────────────────────────────
    stat_defs = [
        (f"Total: {stats['total']}",             "F1F5F9", "334155"),
        (f"✓ Matched: {stats['matched']}",        "D1FAE5", "064E3B"),
        (f"⚠ Mismatched: {stats['mismatched']}",  "FEF3C7", "92400E"),
        (f"Ext Only: {stats['ext_only']}",         "FEE2E2", "991B1B"),
        (f"Sys Only: {stats['sys_only']}",         "EFF6FF", "1E40AF"),
        ("", "FFFFFF", "111111"),
        ("", "FFFFFF", "111111"),
    ]
    for col, (val, bg, fg) in enumerate(stat_defs, 1):
        _style(ws.cell(row=r, column=col, value=val), bg, fg=fg, bold=True, size=10, h="center")
    ws.row_dimensions[r].height = 18
    r += 1

    # Blank spacer
    for col in range(1, COLS + 1):
        ws.cell(row=r, column=col).fill = _fill("FFFFFF")
    ws.row_dimensions[r].height = 6
    r += 1

    # ── Column headers ────────────────────────────────────────────────────────
    hdrs = ["ITEM CODE", "EXT WT (Kgs)", "OPENING STOCK ID",
            "SYS QTY (Kgs)", "DIFFERENCE", "SOLD", "STATUS"]
    for col, h in enumerate(hdrs, 1):
        _style(ws.cell(row=r, column=col, value=h),
               "1A3A6A", fg="FFFFFF", bold=True, size=10,
               h="right" if col in (2, 4, 5) else "left",
               border=_bord)
    ws.row_dimensions[r].height = 18
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    r += 1

    # ── Color maps ────────────────────────────────────────────────────────────
    HDR_COLORS = {
        "match":    ("D1FAE5", "064E3B"),
        "mismatch": ("FEF3C7", "92400E"),
        "ext_only": ("FEE2E2", "991B1B"),
        "sys_only": ("EFF6FF", "1E40AF"),
    }
    DETAIL_BG = {"match": "F9FAFB", "extra": "FFF1F2", "sys": "F0F9FF"}
    ST_LABELS = {
        "match":    ("✓ Match",    "166534"),
        "mismatch": ("⚠ Mismatch", "92400E"),
        "ext_only": ("✗ Ext Only", "991B1B"),
        "sys_only": ("≡ Sys Only", "1E40AF"),
    }

    # ── Data rows ─────────────────────────────────────────────────────────────
    for item in items_data:
        status         = item["status"]
        hdr_bg, hdr_fg = HDR_COLORS.get(status, ("F5F5F5", "111111"))
        st_lbl, st_fg  = ST_LABELS.get(status, ("", "111111"))

        ext_total = f"{item['ext_total']:.3f}" if item["ext_total"] is not None else "—"
        sys_total = f"{item['sys_total']:.3f}" if item["sys_total"] is not None else "—"
        ext_cnt   = str(item["ext_cnt"])         if item["ext_cnt"]   is not None else "—"
        sys_cnt   = str(item["sys_cnt"])         if item["sys_cnt"]   is not None else "—"

        if item["ext_total"] is not None and item["sys_total"] is not None:
            hd       = round(item["ext_total"] - item["sys_total"], 3)
            hdr_diff = f"{hd:+.3f}"
            diff_fg  = "166534" if abs(hd) < 0.001 else ("92400E" if hd > 0 else "1E40AF")
        else:
            hdr_diff, diff_fg = "—", "111111"

        # Item header row
        _write(r,
               [f"▶ {item['item_code']}", ext_total,
                f"{ext_cnt} ext / {sys_cnt} sys entries",
                sys_total, hdr_diff, "", st_lbl],
               hdr_bg,
               fgs    = [hdr_fg, hdr_fg, "777777", hdr_fg, diff_fg, hdr_fg, st_fg],
               bolds  = [True,   True,   False,    True,   True,    False,  True],
               haligns= ["left", "right","left",   "right","right", "left", "left"],
               sizes  = [10,     10,     9,        10,     10,      10,     10],
               row_h=18, top_border=True)
        r += 1

        # Matched / extra weights
        for m in item["matches"]:
            w   = m["ext_weight"]
            sys = m["sys_entry"]
            if sys:
                d        = round(w - sys["qty"], 3)
                dtxt     = f"{d:+.3f}" if abs(d) >= 0.001 else "0.000"
                dfg      = "166534" if abs(d) < 0.001 else "92400E"
                sold_txt = "✦ Sold"      if sys["is_sold"] else "● Available"
                sold_fg  = "991B1B"      if sys["is_sold"] else "065F46"
                sold_bg  = "FECACA"      if sys["is_sold"] else "D1FAE5"

                _write(r,
                       ["", f"{w:.3f}", sys["stock_entry"],
                        f"{sys['qty']:.3f}", dtxt, sold_txt, "✓ Match"],
                       DETAIL_BG["match"],
                       fgs    = ["111111","1E3A5F","1A56DB","111111", dfg,    sold_fg,"166534"],
                       bolds  = [False,   True,   False,   False,    True,   True,   True],
                       haligns= ["left",  "right","left",  "right",  "right","center","left"])
                ws.cell(row=r, column=6).fill = _fill(sold_bg)

            else:
                _write(r,
                       ["", f"{w:.3f}", "— Not in system", "—", f"+{w:.3f}", "—", "✗ Extra"],
                       DETAIL_BG["extra"],
                       fgs    = ["111111","1E3A5F","991B1B","111111","92400E","111111","991B1B"],
                       bolds  = [False,   True,   False,   False,   True,   False,   True],
                       haligns= ["left",  "right","left",  "right", "right","center","left"],
                       italics= [False,   False,  True,    False,   False,  False,   False])
            r += 1

        # System-only entries
        for e in item["unmatched_sys"]:
            sold_txt = "✦ Sold"  if e["is_sold"] else "● Available"
            sold_fg  = "991B1B"  if e["is_sold"] else "065F46"
            sold_bg  = "FECACA"  if e["is_sold"] else "D1FAE5"

            _write(r,
                   ["", "—", e["stock_entry"], f"{e['qty']:.3f}",
                    f"-{e['qty']:.3f}", sold_txt, "≡ Sys Only"],
                   DETAIL_BG["sys"],
                   fgs    = ["111111","111111","1A56DB","111111","1E40AF", sold_fg,"1E40AF"],
                   bolds  = [False,   False,  False,   False,   True,    True,   True],
                   haligns= ["left",  "right","left",  "right", "right", "center","left"])
            ws.cell(row=r, column=6).fill = _fill(sold_bg)
            r += 1

    # ── Grand total row ───────────────────────────────────────────────────────
    r += 1
    total_ext  = sum(i["ext_total"] for i in items_data if i["ext_total"] is not None)
    total_sys  = sum(i["sys_total"] for i in items_data if i["sys_total"] is not None)
    total_diff = round(total_ext - total_sys, 3)

    _write(r,
           ["GRAND TOTAL", f"{total_ext:.3f}", "",
            f"{total_sys:.3f}", f"{total_diff:+.3f}", "", ""],
           "1A3A6A",
           fgs    = ["FFFFFF","FFFFFF","FFFFFF","FFFFFF","FFFFFF","FFFFFF","FFFFFF"],
           bolds  = [True,    True,   False,   True,    True,   False,   False],
           haligns= ["left",  "right","left",  "right", "right","left",  "left"],
           row_h=18, top_border=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return {
        "data":     base64.b64encode(buf.getvalue()).decode(),
        "filename": "stock_diff_report.xlsx",
    }


def _build_diff_data(file_data, filters):
    """Parse external Excel + compare against system entries. Returns (items_data, stats)."""
    import base64, io
    import openpyxl as _xl

    wb = _xl.load_workbook(io.BytesIO(base64.b64decode(file_data)), data_only=True)
    ws = wb.active

    external = {}

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        item_code = str(row[0]).strip()
        if not item_code:
            continue

        nums = []
        for cell in row[1:]:
            if cell is not None:
                try:
                    nums.append(float(cell))
                except (TypeError, ValueError):
                    pass

        if not nums:
            continue

        last = nums[-1]
        if len(nums) >= 2 and last == int(last) and 1 <= int(last) <= len(nums):
            entry_count = int(last)
            weights     = [round(w, 3) for w in nums[:-1] if w > 0]
        else:
            weights     = [round(w, 3) for w in nums if w > 0]
            entry_count = len(weights)

        external[item_code] = {
            "weights":     weights,
            "entry_count": entry_count,
            "total_qty":   round(sum(weights), 3),
        }

    if not external:
        frappe.throw(_("No data rows found in the uploaded Excel file."))

    _, system_rows = execute(filters)
    system       = {}
    current_item = None

    for row in system_rows:
        rtype = row.get("row_type")
        code  = row.get("item_code", "")

        if rtype == "group_header" and code and code != "Grand Total":
            current_item = code
            system[current_item] = {
                "entries":     [],
                "total_qty":   0.0,
                "entry_count": row.get("entry_count", 0),
            }
        elif rtype == "data" and current_item and current_item in system:
            system[current_item]["entries"].append({
                "stock_entry":  row.get("stock_entry", ""),
                "qty":          round(float(row.get("qty") or 0), 3),
                "is_sold":      int(row.get("is_sold", 0) or 0),
                "posting_date": str(row.get("posting_date") or ""),
            })
        elif rtype == "subtotal" and current_item and current_item in system:
            system[current_item]["total_qty"] = round(float(row.get("qty") or 0), 3)

    all_items  = sorted(set(list(external.keys()) + list(system.keys())))
    items_data = []
    stats      = {"total": len(all_items), "matched": 0, "mismatched": 0,
                  "ext_only": 0, "sys_only": 0}

    for item in all_items:
        in_ext  = item in external
        in_sys  = item in system
        weights = external[item]["weights"]  if in_ext else []
        entries = system[item]["entries"]    if in_sys else []

        matches, unmatched_sys = _match_weights(list(weights), list(entries))

        ext_total = external[item]["total_qty"]   if in_ext else None
        sys_total = system[item]["total_qty"]     if in_sys else None
        ext_cnt   = external[item]["entry_count"] if in_ext else None
        sys_cnt   = system[item]["entry_count"]   if in_sys else None

        has_extra    = any(m["sys_entry"] is None for m in matches)
        has_sys_only = bool(unmatched_sys)

        if in_ext and in_sys:
            qty_ok = abs((ext_total or 0) - (sys_total or 0)) < 0.001
            cnt_ok = (ext_cnt or 0) == (sys_cnt or 0)
            if not has_extra and not has_sys_only and qty_ok and cnt_ok:
                status = "match"
                stats["matched"] += 1
            else:
                status = "mismatch"
                stats["mismatched"] += 1
        elif in_ext:
            status = "ext_only"
            stats["ext_only"] += 1
        else:
            status = "sys_only"
            stats["sys_only"] += 1

        items_data.append({
            "item_code":     item,
            "status":        status,
            "ext_total":     ext_total,
            "sys_total":     sys_total,
            "ext_cnt":       ext_cnt,
            "sys_cnt":       sys_cnt,
            "matches":       matches,
            "unmatched_sys": unmatched_sys,
        })

    return items_data, stats


def _match_weights(ext_weights, sys_entries):
    """Pair each external weight to the first system entry within ±0.001 Kgs."""
    used    = [False] * len(sys_entries)
    matches = []
    for w in ext_weights:
        hit = None
        for i, e in enumerate(sys_entries):
            if not used[i] and abs(w - e["qty"]) < 0.001:
                hit     = e
                used[i] = True
                break
        matches.append({"ext_weight": w, "sys_entry": hit})
    unmatched_sys = [e for i, e in enumerate(sys_entries) if not used[i]]
    return matches, unmatched_sys


def _render_diff_html(filters, items_data, stats):
    from datetime import datetime

    company   = filters.get("company", "")
    from_date = str(filters.get("from_date", ""))
    to_date   = str(filters.get("to_date", ""))

    def fmt(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").strftime("%d-%m-%Y")
        except Exception:
            return d

    date_range = f"{fmt(from_date)} to {fmt(to_date)}" if from_date != to_date else fmt(from_date)
    gen_date   = datetime.now().strftime("%d-%m-%Y %H:%M")

    # ── Stats bar ─────────────────────────────────────────────────────────────
    stats_html = f"""
<div class="stats-bar">
  <div class="stat total"><span class="num">{stats['total']}</span><span class="lbl">Total Items</span></div>
  <div class="stat match"><span class="num">{stats['matched']}</span><span class="lbl">&#10003; Matched</span></div>
  <div class="stat mismatch"><span class="num">{stats['mismatched']}</span><span class="lbl">&#9888; Mismatched</span></div>
  <div class="stat ext-only"><span class="num">{stats['ext_only']}</span><span class="lbl">Ext Only</span></div>
  <div class="stat sys-only"><span class="num">{stats['sys_only']}</span><span class="lbl">Sys Only</span></div>
</div>"""

    def sold_badge(is_sold):
        if is_sold:
            return '<span class="badge sold">&#10022; Sold</span>'
        return '<span class="badge avail">&#9679; Available</span>'

    HDR_CLS = {
        "match":    "hdr-match",
        "mismatch": "hdr-mismatch",
        "ext_only": "hdr-ext",
        "sys_only": "hdr-sys",
    }
    ST_LBL = {
        "match":    ("&#10003; Match",    "st-ok"),
        "mismatch": ("&#9888; Mismatch",  "st-warn"),
        "ext_only": ("&#10007; Ext Only", "st-danger"),
        "sys_only": ("&#8801; Sys Only",  "st-info"),
    }

    # ── Per-item rows ─────────────────────────────────────────────────────────
    rows_html = ""

    for item in items_data:
        status    = item["status"]
        code      = item["item_code"]
        lbl, lcls = ST_LBL.get(status, ("", ""))
        hdr_cls   = HDR_CLS.get(status, "")

        ext_total = f"{item['ext_total']:.3f}" if item["ext_total"] is not None else "&#8212;"
        sys_total = f"{item['sys_total']:.3f}" if item["sys_total"] is not None else "&#8212;"
        ext_cnt   = str(item["ext_cnt"])        if item["ext_cnt"]   is not None else "&#8212;"
        sys_cnt   = str(item["sys_cnt"])        if item["sys_cnt"]   is not None else "&#8212;"

        if item["ext_total"] is not None and item["sys_total"] is not None:
            hd = round(item["ext_total"] - item["sys_total"], 3)
            hdr_diff_txt = f"{hd:+.3f}"
            hdr_diff_cls = "ok-num" if abs(hd) < 0.001 else ("hi-num" if hd > 0 else "lo-num")
        else:
            hdr_diff_txt, hdr_diff_cls = "&#8212;", ""

        # Item header row
        rows_html += (
            f'<tr class="item-hdr {hdr_cls}">'
            f'<td class="ic">&#9654; {code}</td>'
            f'<td class="r">{ext_total}</td>'
            f'<td class="cnt-cell">{ext_cnt} ext &nbsp;/&nbsp; {sys_cnt} sys entries</td>'
            f'<td class="r">{sys_total}</td>'
            f'<td class="r {hdr_diff_cls}">{hdr_diff_txt}</td>'
            f'<td></td>'
            f'<td class="st {lcls}">{lbl}</td>'
            f'</tr>\n'
        )

        # Matched + extra external weights
        for m in item["matches"]:
            w   = m["ext_weight"]
            sys = m["sys_entry"]
            if sys:
                d    = round(w - sys["qty"], 3)
                dtxt = f"{d:+.3f}" if abs(d) >= 0.001 else "0.000"
                dcls = "ok-num" if abs(d) < 0.001 else "hi-num"
                rows_html += (
                    f'<tr class="detail-match">'
                    f'<td></td>'
                    f'<td class="r wt">{w:.3f}</td>'
                    f'<td class="se-id">{sys["stock_entry"]}</td>'
                    f'<td class="r">{sys["qty"]:.3f}</td>'
                    f'<td class="r {dcls}">{dtxt}</td>'
                    f'<td>{sold_badge(sys["is_sold"])}</td>'
                    f'<td class="st st-ok">&#10003; Match</td>'
                    f'</tr>\n'
                )
            else:
                rows_html += (
                    f'<tr class="detail-extra">'
                    f'<td></td>'
                    f'<td class="r wt">{w:.3f}</td>'
                    f'<td class="extra-lbl">&#8212; Not in system</td>'
                    f'<td class="r">&#8212;</td>'
                    f'<td class="r hi-num">+{w:.3f}</td>'
                    f'<td>&#8212;</td>'
                    f'<td class="st st-danger">&#10007; Extra</td>'
                    f'</tr>\n'
                )

        # System entries not matched to any external weight
        for e in item["unmatched_sys"]:
            rows_html += (
                f'<tr class="detail-sys">'
                f'<td></td>'
                f'<td class="r">&#8212;</td>'
                f'<td class="se-id">{e["stock_entry"]}</td>'
                f'<td class="r">{e["qty"]:.3f}</td>'
                f'<td class="r lo-num">-{e["qty"]:.3f}</td>'
                f'<td>{sold_badge(e["is_sold"])}</td>'
                f'<td class="st st-info">&#8801; Sys Only</td>'
                f'</tr>\n'
            )

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Stock Difference Report &#8211; {company}</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  @page {{
    size: A4 landscape;
    margin: 10mm 12mm;
  }}

  body {{
    font-family: Arial, Helvetica, sans-serif;
    font-size: 9.5pt;
    color: #111;
    background: #fff;
  }}

  /* ── Header ── */
  .doc-header {{
    text-align: center;
    padding-bottom: 8px;
    border-bottom: 2.5px solid #1a3a6a;
    margin-bottom: 12px;
  }}
  .company-name  {{ font-size:15pt; font-weight:bold; letter-spacing:2px; text-transform:uppercase; color:#1a3a6a; }}
  .report-title  {{ font-size:12pt; margin-top:3px; color:#333; }}
  .report-meta   {{ font-size:8.5pt; color:#666; margin-top:2px; }}

  /* ── Stats bar ── */
  .stats-bar {{
    display: flex;
    border: 1.5px solid #ddd;
    border-radius: 6px;
    overflow: hidden;
    margin-bottom: 14px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  .stat {{ flex:1; text-align:center; padding:8px 4px; }}
  .stat.total    {{ background:#f1f5f9; }}
  .stat.match    {{ background:#d1fae5; }}
  .stat.mismatch {{ background:#fef3c7; }}
  .stat.ext-only {{ background:#fee2e2; }}
  .stat.sys-only {{ background:#eff6ff; }}
  .stat .num {{ display:block; font-size:18pt; font-weight:bold; line-height:1.1; }}
  .stat.total    .num {{ color:#334155; }}
  .stat.match    .num {{ color:#064e3b; }}
  .stat.mismatch .num {{ color:#92400e; }}
  .stat.ext-only .num {{ color:#991b1b; }}
  .stat.sys-only .num {{ color:#1e40af; }}
  .stat .lbl {{ font-size:8pt; color:#555; margin-top:2px; display:block; }}

  /* ── Table ── */
  table {{ width:100%; border-collapse:collapse; table-layout:fixed; }}

  col.c-item {{ width:14%; }}
  col.c-ewt  {{ width:11%; }}
  col.c-se   {{ width:22%; }}
  col.c-sqty {{ width:11%; }}
  col.c-diff {{ width:10%; }}
  col.c-sold {{ width:13%; }}
  col.c-stat {{ width:11%; }}

  thead th {{
    background: #1a3a6a;
    color: #fff;
    border: 1px solid #2a4a8a;
    padding: 5px 6px;
    font-size: 9.5pt;
    font-weight: bold;
    text-align: left;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  thead th.r {{ text-align:right; }}

  tbody td {{
    border: 1px solid #d0d0d0;
    padding: 3px 6px;
    font-size: 9.5pt;
    vertical-align: middle;
    overflow: hidden;
    white-space: nowrap;
    text-overflow: ellipsis;
  }}
  tbody td.r {{ text-align:right; }}

  /* ── Item header rows ── */
  tr.item-hdr td {{
    font-weight: bold;
    font-size: 10pt;
    padding: 5px 6px;
    border-top: 2px solid #888;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  tr.hdr-match    td {{ background:#d1fae5; color:#064e3b; border-color:#4ade80; }}
  tr.hdr-mismatch td {{ background:#fef3c7; color:#92400e; border-color:#fbbf24; }}
  tr.hdr-ext      td {{ background:#fee2e2; color:#991b1b; border-color:#fca5a5; }}
  tr.hdr-sys      td {{ background:#eff6ff; color:#1e40af; border-color:#93c5fd; }}

  /* ── Detail rows ── */
  tr.detail-match td {{ background:#f9fafb; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
  tr.detail-extra td {{ background:#fff1f2; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
  tr.detail-sys   td {{ background:#f0f9ff; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}

  td.ic        {{ font-weight:700; font-size:10pt; }}
  td.wt        {{ font-weight:600; color:#1e3a5f; }}
  td.se-id     {{ font-family:monospace; font-size:8.5pt; color:#1a56db; }}
  td.extra-lbl {{ font-style:italic; color:#991b1b; font-size:8.5pt; }}
  td.cnt-cell  {{ font-size:8.5pt; color:#555; }}

  /* ── Diff colours ── */
  td.ok-num {{ color:#166534; font-weight:bold; }}
  td.hi-num {{ color:#92400e; font-weight:bold; }}
  td.lo-num {{ color:#1e40af; font-weight:bold; }}

  /* ── Status ── */
  td.st          {{ font-weight:700; font-size:9pt; }}
  td.st-ok       {{ color:#166534; }}
  td.st-warn     {{ color:#92400e; }}
  td.st-danger   {{ color:#991b1b; }}
  td.st-info     {{ color:#1e40af; }}

  /* ── Sold / Available badges ── */
  .badge {{
    display: inline-block;
    border-radius: 10px;
    font-size: 8pt;
    font-weight: 700;
    padding: 1px 9px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  .badge.sold  {{ background:#fecaca; color:#991b1b; border:1px solid #f87171; }}
  .badge.avail {{ background:#d1fae5; color:#065f46; border:1px solid #6ee7b7; }}

  /* ── Legend ── */
  .legend {{
    margin-top: 10px;
    font-size: 8.5pt;
    color: #555;
    display: flex;
    gap: 18px;
    flex-wrap: wrap;
    align-items: center;
  }}
  .legend-item {{ display:flex; align-items:center; gap:5px; }}
  .legend-dot  {{ width:12px; height:12px; border-radius:2px; flex-shrink:0; }}

  /* ── Print ── */
  @media print {{
    thead {{ display:table-header-group; }}
    .no-print {{ display:none; }}
    tr.item-hdr {{ page-break-before:auto; page-break-after:avoid; }}
  }}

  .print-btn-bar {{ text-align:right; margin-bottom:8px; }}
  .print-btn {{
    padding: 6px 18px;
    background: #1a73e8;
    color: #fff;
    border: none;
    border-radius: 4px;
    font-size: 10pt;
    cursor: pointer;
  }}
  .print-btn:hover {{ background:#1558b0; }}
</style>
</head>
<body>

<div class="print-btn-bar no-print">
  <button class="print-btn" onclick="window.print()">&#128438; Print / Save as PDF</button>
</div>

<div class="doc-header">
  <div class="company-name">{company}</div>
  <div class="report-title">Stock Difference Report &nbsp;&#8212;&nbsp; {date_range}</div>
  <div class="report-meta">Generated: {gen_date} &nbsp;|&nbsp; External Excel vs Opening Stock Audit</div>
</div>

{stats_html}

<table>
  <colgroup>
    <col class="c-item">
    <col class="c-ewt">
    <col class="c-se">
    <col class="c-sqty">
    <col class="c-diff">
    <col class="c-sold">
    <col class="c-stat">
  </colgroup>
  <thead>
    <tr>
      <th>ITEM CODE</th>
      <th class="r">EXT WT (Kgs)</th>
      <th>OPENING STOCK ID</th>
      <th class="r">SYS QTY (Kgs)</th>
      <th class="r">DIFFERENCE</th>
      <th>SOLD</th>
      <th>STATUS</th>
    </tr>
  </thead>
  <tbody>
    {rows_html}
  </tbody>
</table>

<div class="legend no-print">
  <div class="legend-item">
    <div class="legend-dot" style="background:#d1fae5;border:1px solid #4ade80;"></div>Item fully matched
  </div>
  <div class="legend-item">
    <div class="legend-dot" style="background:#fef3c7;border:1px solid #fbbf24;"></div>Item has mismatch
  </div>
  <div class="legend-item">
    <div class="legend-dot" style="background:#fff1f2;border:1px solid #fca5a5;"></div>Weight extra in external
  </div>
  <div class="legend-item">
    <div class="legend-dot" style="background:#f0f9ff;border:1px solid #93c5fd;"></div>Entry only in system
  </div>
  <span style="margin-left:8px;color:#666;">Difference = Ext Weight &#8722; Sys Qty</span>
</div>

</body>
</html>"""


@frappe.whitelist()
def get_print_html(filters):
    if isinstance(filters, str):
        filters = json.loads(filters)
    _, data = execute(filters)
    return _render_html(filters, data)


def _render_html(filters, data):
    from datetime import datetime

    company    = filters.get("company", "")
    from_date  = str(filters.get("from_date", ""))
    to_date    = str(filters.get("to_date", ""))

    def fmt(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").strftime("%d-%m-%Y")
        except Exception:
            return d

    date_range = f"{fmt(from_date)} to {fmt(to_date)}" if from_date != to_date else fmt(from_date)
    rows_html  = _build_table_rows(data)

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Opening Stock Audit – {company}</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  @page {{
    size: A4 portrait;
    margin: 14mm 12mm 16mm 12mm;
  }}

  body {{
    font-family: Arial, Helvetica, sans-serif;
    font-size: 10pt;
    color: #111;
    background: #fff;
  }}

  /* ── Header ── */
  .doc-header {{
    text-align: center;
    padding-bottom: 6px;
    border-bottom: 2px solid #222;
    margin-bottom: 6px;
  }}
  .doc-header .company-name {{
    font-size: 16pt;
    font-weight: bold;
    letter-spacing: 3px;
    text-transform: uppercase;
  }}
  .doc-header .report-title {{
    font-size: 11pt;
    margin-top: 2px;
    color: #333;
  }}

  /* ── Table ── */
  table {{
    width: 100%;
    border-collapse: collapse;
    table-layout: fixed;
  }}

  col.c-code  {{ width: 15%; }}
  col.c-grp   {{ width: 13%; }}
  col.c-se    {{ width: 17%; }}
  col.c-date  {{ width: 10%; }}
  col.c-qty   {{ width: 11%; }}
  col.c-wh    {{ width: 26%; }}
  col.c-sold  {{ width:  8%; }}

  thead th {{
    background: #1a3a6a;
    color: #fff;
    border: 1px solid #2a4a8a;
    padding: 5px 6px;
    font-size: 10pt;
    font-weight: bold;
    text-align: left;
  }}
  thead th.r {{ text-align: right; }}

  tbody td {{
    border: 1px solid #ccc;
    padding: 3px 6px;
    font-size: 10pt;
    vertical-align: middle;
    overflow: hidden;
    white-space: nowrap;
    text-overflow: ellipsis;
  }}
  tbody td.r {{ text-align: right; }}

  /* ── Single entry group header — green ── */
  tr.grp-ok td {{
    background: #d1fae5;
    border: 1.5px solid #6ee7b7;
    font-weight: bold;
    font-size: 10.5pt;
    color: #064e3b;
    padding: 4px 6px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}

  /* ── Multi entry group header — orange warning ── */
  tr.grp-warn td {{
    background: #fef3c7;
    border: 1.5px solid #fbbf24;
    font-weight: bold;
    font-size: 10.5pt;
    color: #92400e;
    padding: 4px 6px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  tr.grp-warn td.badge-cell::after {{
    content: "⚠ " attr(data-count) " entries";
    display: inline-block;
    background: #fbbf24;
    color: #7c2d12;
    border-radius: 8px;
    font-size: 8pt;
    font-weight: 700;
    padding: 1px 7px;
    margin-left: 8px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}

  /* ── Data rows ── */
  tr.data-ok td {{
    background: #f0fdf4;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  tr.data-warn td {{
    background: #fffbeb;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}

  /* ── Subtotal ── */
  tr.subtotal-ok td {{
    background: #dcfce7;
    border-top: 2px solid #4ade80;
    border-bottom: 2px solid #4ade80;
    font-weight: bold;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  tr.subtotal-warn td {{
    background: #fde68a;
    border-top: 2px solid #f59e0b;
    border-bottom: 2px solid #f59e0b;
    font-weight: bold;
    color: #92400e;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}

  /* ── Grand total ── */
  tr.grand-total td {{
    background: #1a3a6a;
    color: #fff;
    border: 2px solid #1a3a6a;
    font-weight: bold;
    font-size: 11pt;
    padding: 5px 6px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}

  /* ── Sold overlay — overrides the row colour for sold items ── */
  tr.sold td {{
    background: #fee2e2 !important;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  .sold-badge {{
    display: inline-block;
    background: #fecaca;
    color: #991b1b;
    border: 1px solid #f87171;
    border-radius: 8px;
    font-size: 8pt;
    font-weight: 700;
    padding: 1px 6px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}

  /* ── Print tweaks ── */
  @media print {{
    thead {{ display: table-header-group; }}
    tr.grp-ok, tr.grp-warn {{ page-break-after: avoid; }}
    tr.subtotal-ok, tr.subtotal-warn {{ page-break-before: avoid; }}
    .no-print {{ display: none; }}
  }}

  .print-btn-bar {{ text-align: right; margin-bottom: 8px; }}
  .print-btn {{
    padding: 7px 20px;
    background: #1a73e8;
    color: #fff;
    border: none;
    border-radius: 4px;
    font-size: 10pt;
    cursor: pointer;
  }}
  .print-btn:hover {{ background: #1558b0; }}
</style>
</head>
<body>

<div class="print-btn-bar no-print">
  <button class="print-btn" onclick="window.print()">&#128438; Print / Save as PDF</button>
</div>

<div class="doc-header">
  <div class="company-name">{company}</div>
  <div class="report-title">Opening Stock Audit &nbsp;—&nbsp; {date_range}</div>
</div>

<table>
  <colgroup>
    <col class="c-code">
    <col class="c-grp">
    <col class="c-se">
    <col class="c-date">
    <col class="c-qty">
    <col class="c-wh">
    <col class="c-sold">
  </colgroup>
  <thead>
    <tr>
      <th>ITEM CODE</th>
      <th>ITEM GROUP</th>
      <th>STOCK ENTRY</th>
      <th>DATE</th>
      <th class="r">QTY (KGS)</th>
      <th>WAREHOUSE</th>
      <th style="text-align:center;">SOLD</th>
    </tr>
  </thead>
  <tbody>
    {rows_html}
  </tbody>
</table>

</body>
</html>"""


def _build_table_rows(data):
    from datetime import datetime

    def fdate(d):
        if not d:
            return ""
        try:
            return datetime.strptime(str(d), "%Y-%m-%d").strftime("%d-%m-%Y")
        except Exception:
            return str(d)

    html = ""
    for row in data:
        rtype = row.get("row_type", "data")
        multi = (row.get("entry_count") or 1) > 1

        is_sold   = row.get("is_sold")
        sold_cls  = " sold" if is_sold else ""
        sold_cell = '<td style="text-align:center;"><span class="sold-badge">✦ Yes</span></td>' if is_sold else "<td></td>"

        if rtype == "group_header":
            tr_cls = ("grp-warn" if multi else "grp-ok") + sold_cls
            badge  = (f'<span style="background:#f59e0b;color:#7c2d12;border-radius:8px;'
                      f'font-size:8pt;font-weight:700;padding:1px 7px;margin-left:8px;'
                      f'-webkit-print-color-adjust:exact;print-color-adjust:exact;">'
                      f'⚠ {row["entry_count"]} entries</span>') if multi else ""
            html += (
                f'<tr class="{tr_cls}">'
                f'<td><strong>{row.get("item_code","")}</strong></td>'
                f'<td>{row.get("item_group","")}{badge}</td>'
                f'<td></td><td></td><td></td><td></td>'
                f'{sold_cell}'
                f'</tr>\n'
            )

        elif rtype == "data":
            tr_cls = ("data-warn" if multi else "data-ok") + sold_cls
            qty    = f'{row.get("qty",0):.3f}' if row.get("qty") is not None else ""
            html += (
                f'<tr class="{tr_cls}">'
                f'<td></td>'
                f'<td></td>'
                f'<td>{row.get("stock_entry","")}</td>'
                f'<td>{fdate(row.get("posting_date"))}</td>'
                f'<td class="r"><strong>{qty}</strong></td>'
                f'<td>{row.get("warehouse","")}</td>'
                f'{sold_cell}'
                f'</tr>\n'
            )

        elif rtype == "subtotal":
            tr_cls = ("subtotal-warn" if multi else "subtotal-ok") + sold_cls
            qty    = f'{row.get("qty",0):.3f}' if row.get("qty") is not None else ""
            label  = row.get("stock_entry", "")
            html += (
                f'<tr class="{tr_cls}">'
                f'<td colspan="3" style="text-align:right;padding-right:8px">{label}</td>'
                f'<td></td>'
                f'<td class="r">{qty}</td>'
                f'<td></td>'
                f'<td></td>'
                f'</tr>\n'
            )

        elif rtype == "grand_total":
            qty = f'{row.get("qty",0):.3f}' if row.get("qty") is not None else ""
            html += (
                f'<tr class="grand-total">'
                f'<td><strong>{row.get("item_code","")}</strong></td>'
                f'<td colspan="3">{row.get("item_group","")}</td>'
                f'<td class="r"><strong>{qty}</strong></td>'
                f'<td></td>'
                f'<td></td>'
                f'</tr>\n'
            )

    return html

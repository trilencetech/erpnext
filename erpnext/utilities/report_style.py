"""Shared styling helpers for branded internal register/dashboard reports.

Provides a consistent color palette, openpyxl cell-styling helpers, and an
HTML print shell (header + stats bar + table) reused across the
"*_register" reports (Purchase Receipt, Purchase Invoice, Delivery Note,
Sales Invoice). Status/condition coloring is expressed via a small set of
"kinds": ok (green), warn (amber), error (red), info (blue), neutral (gray).
"""

from datetime import datetime


# ── Color palette ────────────────────────────────────────────────────────────

COLORS = {
    "header_bg": "1A3A6A",
    "header_fg": "FFFFFF",
    "neutral_bg": "F1F5F9",
    "neutral_fg": "334155",
    "border": "D0D0D0",
}

# kind -> (background, foreground, border) hex colors, used for both
# HTML badges/rows and Excel cell styling.
KIND_COLORS = {
    "ok":      ("D1FAE5", "064E3B", "6EE7B7"),
    "warn":    ("FEF3C7", "92400E", "FBBF24"),
    "error":   ("FEE2E2", "991B1B", "FCA5A5"),
    "info":    ("EFF6FF", "1E40AF", "93C5FD"),
    "neutral": ("F1F5F9", "334155", "CBD5E1"),
}


# ── openpyxl helpers ──────────────────────────────────────────────────────────

def cell_fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=hex_color)


def cell_font(bold=False, color="111111", size=10, italic=False):
    from openpyxl.styles import Font
    return Font(bold=bold, color=color, name="Arial", size=size, italic=italic)


def cell_align(h="left", v="center"):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=False)


def _build_borders():
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color=COLORS["border"])
    medium = Side(style="medium", color="888888")
    return (
        Border(left=thin, right=thin, top=thin, bottom=thin),
        Border(left=thin, right=thin, top=medium, bottom=thin),
    )


BORDER_THIN, BORDER_TOP_MEDIUM = _build_borders()


def style_cell(cell, bg, fg="111111", bold=False, size=10, h="left", italic=False, border=None):
    cell.fill = cell_fill(bg)
    cell.font = cell_font(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = cell_align(h)
    cell.border = border or BORDER_THIN


def merge_and_style(ws, row, last_col, value, bg, fg="111111", bold=False, size=10, h="center"):
    """Merge columns 1..last_col of `row` and write `value` into the first cell."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=value)
    style_cell(c, bg, fg=fg, bold=bold, size=size, h=h)
    for col in range(2, last_col + 1):
        ws.cell(row=row, column=col).fill = cell_fill(bg)


def write_row(ws, row, values, bg, fgs=None, bolds=None, haligns=None, sizes=None, italics=None, row_h=17, top_border=False):
    """Write `values` across columns 1..len(values) of `row` with per-column styling."""
    border = BORDER_TOP_MEDIUM if top_border else BORDER_THIN
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        style_cell(
            c, bg,
            fg=fgs[col - 1] if fgs else "111111",
            bold=bolds[col - 1] if bolds else False,
            h=haligns[col - 1] if haligns else "left",
            size=sizes[col - 1] if sizes else 10,
            italic=italics[col - 1] if italics else False,
            border=border,
        )
    ws.row_dimensions[row].height = row_h


def badge_xlsx(ws, row, col, label, kind="neutral"):
    """Write a status-badge-like cell (colored fill + bold colored text)."""
    bg, fg, _border = KIND_COLORS.get(kind, KIND_COLORS["neutral"])
    c = ws.cell(row=row, column=col, value=label)
    style_cell(c, bg, fg=fg, bold=True, size=9, h="center")
    return c


# ── HTML helpers ──────────────────────────────────────────────────────────────

PRINT_CSS = """
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  @page {
    size: A4 landscape;
    margin: 12mm 10mm 14mm 10mm;
  }

  body {
    font-family: Arial, Helvetica, sans-serif;
    font-size: 10pt;
    color: #111;
    background: #fff;
  }

  /* ── Header ── */
  .doc-header {
    text-align: center;
    padding-bottom: 6px;
    border-bottom: 2px solid #222;
    margin-bottom: 8px;
  }
  .doc-header .company-name {
    font-size: 16pt;
    font-weight: bold;
    letter-spacing: 3px;
    text-transform: uppercase;
  }
  .doc-header .report-title {
    font-size: 11pt;
    margin-top: 2px;
    color: #333;
  }

  /* ── Stats bar ── */
  .stats-bar {
    display: flex;
    border: 1.5px solid #ddd;
    border-radius: 6px;
    overflow: hidden;
    margin-bottom: 10px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }
  .stat { flex: 1; text-align: center; padding: 8px 4px; }
  .stat .num { display: block; font-size: 18pt; font-weight: bold; line-height: 1.1; }
  .stat .lbl { font-size: 8pt; color: #555; margin-top: 2px; display: block; }
  .stat.ok      { background: #d1fae5; } .stat.ok      .num { color: #064e3b; }
  .stat.warn    { background: #fef3c7; } .stat.warn    .num { color: #92400e; }
  .stat.error   { background: #fee2e2; } .stat.error   .num { color: #991b1b; }
  .stat.info    { background: #eff6ff; } .stat.info    .num { color: #1e40af; }
  .stat.neutral { background: #f1f5f9; } .stat.neutral .num { color: #334155; }

  /* ── Table ── */
  table {
    width: 100%;
    border-collapse: collapse;
  }

  thead th {
    background: #1a3a6a;
    color: #fff;
    border: 1px solid #2a4a8a;
    padding: 5px 6px;
    font-size: 10pt;
    font-weight: bold;
    text-align: left;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }
  thead th.r { text-align: right; }

  tbody td {
    border: 1px solid #ccc;
    padding: 3px 6px;
    font-size: 10pt;
    vertical-align: middle;
    overflow: hidden;
    white-space: nowrap;
    text-overflow: ellipsis;
  }
  tbody td.r { text-align: right; }

  /* ── Row "kind" coloring (group headers / subtotals) ── */
  tr.grp-ok td,    tr.sub-ok td      { background: #d1fae5; color: #064e3b; }
  tr.grp-warn td,  tr.sub-warn td    { background: #fef3c7; color: #92400e; }
  tr.grp-error td, tr.sub-error td   { background: #fee2e2; color: #991b1b; }
  tr.grp-info td,  tr.sub-info td    { background: #eff6ff; color: #1e40af; }
  tr.grp-neutral td, tr.sub-neutral td { background: #f1f5f9; color: #334155; }

  tr.grp-ok td, tr.grp-warn td, tr.grp-error td, tr.grp-info td, tr.grp-neutral td {
    border: 1.5px solid #999;
    font-weight: bold;
    font-size: 10.5pt;
    padding: 4px 6px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }
  tr.sub-ok td, tr.sub-warn td, tr.sub-error td, tr.sub-info td, tr.sub-neutral td {
    font-weight: bold;
    border-top: 2px solid #999;
    border-bottom: 2px solid #999;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }

  /* ── Grand total ── */
  tr.grand-total td {
    background: #1a3a6a;
    color: #fff;
    border: 2px solid #1a3a6a;
    font-weight: bold;
    font-size: 11pt;
    padding: 5px 6px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }

  /* ── Badges ── */
  .badge {
    display: inline-block;
    border-radius: 10px;
    font-size: 8pt;
    font-weight: 700;
    padding: 1px 9px;
    margin-left: 6px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }
  .badge.ok      { background: #d1fae5; color: #065f46; border: 1px solid #6ee7b7; }
  .badge.warn    { background: #fef3c7; color: #92400e; border: 1px solid #fbbf24; }
  .badge.error   { background: #fecaca; color: #991b1b; border: 1px solid #f87171; }
  .badge.info    { background: #dbeafe; color: #1e40af; border: 1px solid #93c5fd; }
  .badge.neutral { background: #e2e8f0; color: #334155; border: 1px solid #cbd5e1; }

  /* ── Print tweaks ── */
  @media print {
    thead { display: table-header-group; }
    tr.grp-ok, tr.grp-warn, tr.grp-error, tr.grp-info, tr.grp-neutral { page-break-after: avoid; }
    tr.sub-ok, tr.sub-warn, tr.sub-error, tr.sub-info, tr.sub-neutral { page-break-before: avoid; }
    .no-print { display: none; }
  }

  .print-btn-bar { text-align: right; margin-bottom: 8px; }
  .print-btn {
    padding: 7px 20px;
    background: #1a73e8;
    color: #fff;
    border: none;
    border-radius: 4px;
    font-size: 10pt;
    cursor: pointer;
  }
  .print-btn:hover { background: #1558b0; }
"""


def badge_html(label, kind="neutral"):
    return f'<span class="badge {kind}">{label}</span>'


def render_stats_bar(stats):
    """`stats` is a list of {"label": str, "value": int|str, "kind": "ok"|"warn"|"error"|"info"|"neutral"}."""
    boxes = "".join(
        f'<div class="stat {s["kind"]}">'
        f'<span class="num">{s["value"]}</span>'
        f'<span class="lbl">{s["label"]}</span>'
        f'</div>'
        for s in stats
    )
    return f'<div class="stats-bar">{boxes}</div>'


def render_print_shell(title, company, subtitle, stats_html, table_html):
    """Assemble the full printable HTML document."""
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>{title} – {company}</title>
<style>
{PRINT_CSS}
</style>
</head>
<body>

<div class="print-btn-bar no-print">
  <button class="print-btn" onclick="window.print()">&#128438; Print / Save as PDF</button>
</div>

<div class="doc-header">
  <div class="company-name">{company}</div>
  <div class="report-title">{title} &nbsp;—&nbsp; {subtitle}</div>
</div>

{stats_html}

{table_html}

</body>
</html>"""


def fmt_date(d):
    if not d:
        return ""
    try:
        return datetime.strptime(str(d), "%Y-%m-%d").strftime("%d-%m-%Y")
    except Exception:
        return str(d)


def date_range_label(from_date, to_date):
    from_date, to_date = str(from_date or ""), str(to_date or "")
    if from_date and from_date != to_date:
        return f"{fmt_date(from_date)} to {fmt_date(to_date)}"
    return fmt_date(from_date or to_date)
